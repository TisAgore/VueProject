import csv
from pathlib import Path
from typing import Any

from auditmate.models.attachments import Attachment, ParsedAttachment, create_attachment
from auditmate.tools.spreadsheets.financial_extractor import extract_financial_metrics


class CSVParserTool:
    name = "parse_csv"
    description = "Читает CSV-файлы и извлекает вероятные финансовые метрики."

    def execute(self, input_data: Attachment | str | Path) -> ParsedAttachment:
        """Парсит CSV-файл в строки таблицы и текстовое представление."""

        attachment = _coerce_attachment(input_data)
        with attachment.path.open(encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))

        return _build_table_attachment(attachment, rows)


class XLSXParserTool:
    name = "parse_xlsx"
    description = "Читает XLSX/XLS-файлы и извлекает вероятные финансовые метрики."

    def execute(self, input_data: Attachment | str | Path) -> ParsedAttachment:
        """Парсит Excel-файл через openpyxl, если табличная зависимость установлена."""
        from openpyxl import load_workbook

        attachment = _coerce_attachment(input_data)
        rows: list[dict[str, Any]] = []
        workbook = load_workbook(attachment.path, data_only=True, read_only=True)
        try:
            for sheet in workbook.worksheets:
                values = list(sheet.iter_rows(values_only=True))
                if not values:
                    continue

                headers = [
                    str(header).strip() if header not in (None, "") else f"column_{index + 1}"
                    for index, header in enumerate(values[0])
                ]
                for raw_row in values[1:]:
                    row = {
                        headers[index]: value
                        for index, value in enumerate(raw_row)
                        if index < len(headers)
                    }
                    row["_sheet"] = sheet.title
                    rows.append(row)
        finally:
            workbook.close()

        return _build_table_attachment(attachment, rows)


def _build_table_attachment(attachment: Attachment, rows: list[dict[str, Any]]) -> ParsedAttachment:
    """Создаёт ParsedAttachment для таблицы и добавляет финансовые метрики."""

    extraction = extract_financial_metrics(rows)
    text = _rows_to_text(rows)
    return ParsedAttachment(
        attachment=attachment,
        text=text,
        data={
            "rows": rows,
            "financial_metrics": extraction.metrics,
            "financial_evidence": extraction.evidence,
        },
        metadata={
            "source": str(attachment.path),
            "kind": attachment.kind,
            "rows": len(rows),
            "characters": len(text),
        },
    )


def _rows_to_text(rows: list[dict[str, Any]], max_rows: int = 80) -> str:
    """Преобразует первые строки таблицы в компактный текст для LLM-контекста."""

    lines = []
    for row in rows[:max_rows]:
        cells = [f"{key}: {value}" for key, value in row.items() if value not in (None, "")]
        if cells:
            lines.append(" | ".join(cells))

    if len(rows) > max_rows:
        lines.append(f"... ещё строк: {len(rows) - max_rows}")

    return "\n".join(lines)


def _coerce_attachment(input_data: Attachment | str | Path) -> Attachment:
    """Нормализует вход парсера в объект Attachment."""

    if isinstance(input_data, Attachment):
        return input_data

    return create_attachment(input_data)
